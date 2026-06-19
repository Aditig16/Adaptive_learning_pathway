import os
import sys
import json
import re
import time
import random
import difflib
import openpyxl
from dotenv import load_dotenv
from google import genai
from google.genai import types
#from fetch_content import content_generation


SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(SCRIPT_DIR, "data")

USER_FILE = os.path.join(DATA_DIR, "HCAI_User_dataset.xlsx")
ROLE_FILE = os.path.join(DATA_DIR, "IT_Roles_Dataset.xlsx")

STUDY_MATERIAL = os.path.join(DATA_DIR, "Data_Scientist_YouTube_Resources.xlsx")


MODEL_CANDIDATES = [
    "gemini-2.5-flash",
    "gemini-1.5-flash",
    "gemini-1.5-flash-latest",
    "gemini-1.5-pro",
]
SKILL_ALIASES = {
    "python programming": "Python",
    "python": "Python",
    "js": "JavaScript",
    "javascript": "JavaScript",
    "sql database": "SQL",
    "jupyter notebook": "Jupyter",
    "structured query language": "SQL",
    "data visualisation": "Data Visualization",
    "data visualization": "Data Visualization",
    "powerbi": "Power BI",
    "power bi": "Power BI",
    "version control (git)": "Git",
    "git": "Git"
    
}
NOISE = {
    "student", "bachelor’s", "master’s", "phd",
    "no experience(0)", "no experience",
    "get first job", "internship preparation",
    "working professional", "career switch",
    "outlook notes:", "column", "timestamp"
}

def clean(v):
    if v is None:
        return None
    v = str(v).strip()
    if v.lower() in ("none", "nan", "", "null"):
        return None
    return v


def split_skill_cell(cell):
    if not cell:
        return []
    text = str(cell)

    parts = [p.strip() for p in text.split(";") if p.strip()]
    return parts


def normalize_skill(s):
    if not s:
        return None

    s = str(s).strip()
    low = s.lower()

    if low in NOISE:
        return None

    if low in SKILL_ALIASES:
        return SKILL_ALIASES[low]

    if any(char.isdigit() for char in s) and "-" in s:
        return None

    if len(s.split()) > 6:
        return None

    if len(s) <= 2:
        return None

    return s


def flatten_user_skills(row_dict):
    skills = []

    for k, v in row_dict.items():
        if not v:
            continue

        if "skill" in k.lower() or "tool" in k.lower():
            parts = split_skill_cell(v)
            if parts:
                skills.extend(parts)
            else:
                skills.append(v)

    cleaned = []
    for s in skills:
        s = normalize_skill(s)
        if s:
            cleaned.append(s)

    return list(dict.fromkeys(cleaned))


def load_users():
    wb = openpyxl.load_workbook(USER_FILE)
    ws = wb.active

    headers = [clean(h) for h in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]

    users = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        row = [clean(x) for x in row]
        row_dict = {headers[i]: row[i] for i in range(min(len(headers), len(row)))}

        users.append({
            "id": len(users) + 1,
            "name": row_dict.get("Full Name"),
            "target_role": row_dict.get("What IT role are you targeting?"),
            "timelines":row_dict.get("What is your target timeline?"),
            "raw_row": row_dict
        })
    return users


def load_roles():
    wb = openpyxl.load_workbook(ROLE_FILE)
    ws = wb.active

    headers = [clean(h) for h in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]

    def get(row_dict, name):
        for k, v in row_dict.items():
            if k and name.lower() in k.lower():
                return v
        return None

    roles = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        row = [clean(x) for x in row]
        row_dict = {headers[i]: row[i] for i in range(min(len(headers), len(row)))}

        role = get(row_dict, "Role")
        if not role:
            continue

        roles[role] = {
            "required": split_skill_cell(get(row_dict, "Core / Required Skills")),
            "preferred": split_skill_cell(get(row_dict, "Preferred Skills")),
            "tools": split_skill_cell(get(row_dict, "Tools")),
        }

    return roles


def match_role(user_role, roles):
    if not user_role:
        return None

    user_role = clean(user_role)

    if user_role in roles:
        return user_role

    match = difflib.get_close_matches(
        user_role,
        list(roles.keys()),
        n=1,
        cutoff=0.4
    )

    return match[0] if match else None


def parse_experience(exp):
    if not exp:
        return 0

    exp = str(exp).lower()

    if "fresher" in exp or "no experience" in exp or "0" in exp:
        return 0
    if "1" in exp or "1-2" in exp or "1–2" in exp:
        return 1
    if "2" in exp or "2-4" in exp or "2–4" in exp:
        return 2
    if "3" in exp or "3-5" in exp or "3–5" in exp:
        return 3

    return 1


def align_skills(user_skills, role_data):
    if not role_data:
        return user_skills

    role_skills = set(
        (role_data.get("required") or []) +
        (role_data.get("preferred") or []) +
        (role_data.get("tools") or [])
    )

    if not role_skills:
        return user_skills

    aligned = []

    for s in user_skills:
        for r in role_skills:
            if not s or not r:
                continue

            s_tokens = set(s.lower().split())
            r_tokens = set(r.lower().split())

            if len(s_tokens.intersection(r_tokens)) > 0:
                aligned.append(s)
                break

    return aligned if aligned else user_skills


def filter_relevant_skills(user_skills, role_data):
    if not role_data:
        return user_skills

    role_skills = set(
        (role_data.get("required") or []) +
        (role_data.get("preferred") or []) +
        (role_data.get("tools") or [])
    )

    filtered = []

    for s in user_skills:
        for r in role_skills:
            if s and r and (s.lower() in r.lower() or r.lower() in s.lower()):
                filtered.append(s)
                break

    return list(dict.fromkeys(filtered))


def build_difficulty_map(experience_level):
    if experience_level == 0:
        return ["easy"]          # only easy
    elif experience_level == 1:
        return ["easy", "medium"]
    elif experience_level == 2:
        return ["medium", "hard"]
    else:
        return ["easy", "medium", "hard"]

def call_gemini(client, prompt, retries=5):
    for model in MODEL_CANDIDATES:
        for i in range(retries):
            try:
                resp = client.models.generate_content(
                    model=model,
                    contents=prompt,
                    config=types.GenerateContentConfig(
                        response_mime_type="application/json"
                    )
                )
                return json.loads(resp.text)

            except Exception as e:
                print("GEMINI ERROR:", e)
                msg = str(e)

                if "429" in msg or "503" in msg or "overloaded" in msg:
                    time.sleep((2 ** i) + random.random())
                    continue

                if "404" in msg or "not found" in msg:
                    break

                break

    return None


def generate_mcqs(client, payload, role):
    prompt = f"""
Generate MCQs for skill assessment. 1 question per skill

ROLE: {role}

SKILLS (ONLY USE THESE):
{payload["skills"]}

DIFFICULTY DISTRIBUTION:
{payload["difficulty"]}

RULES:
- Only use given skills
- Respect difficulty distribution
- No external topics

Return JSON:
{{
  "questions": [
    {{
      "skill": "",
      "difficulty": "",
      "question": "",
      "options": {{"A":"","B":"","C":"","D":""}},
      "correct answer": ""
      "user answer": ""
    }}
  ]
}}
"""
    return call_gemini(client, prompt)


def skill_gap(result):
    score_per_skill={}
    for res in result["questions"]:
        skill= res["skill"]
        correct_answer=res["correct answer"]
        user_answer=res["user answer"]
        if skill not in score_per_skill:
            score_per_skill[skill]={
                "correct":0,
                "total_questions":0
            }
        score_per_skill[skill]["total_questions"] += 1
        if user_answer.lower()==correct_answer.lower():
            score_per_skill[skill]["correct"]+=1
    
    return calculate_percentage(score_per_skill)        
        
def calculate_percentage(scores):
    percent={}
    for skill, score in scores.items():
        correct = score["correct"]
        total = score["total_questions"]

        percentage = round((correct/total)*100,2)
        percent[skill]=percentage
    
    return percent
def expand_skill(skill):
    if not skill:
        return []

    skill = str(skill).strip()

    skill = skill.replace("’", "'")
    match = re.match(r"^(.*?)\((.*?)\)\s*$", skill)

    if match:
        parent = match.group(1).strip()
        nested = [x.strip() for x in match.group(2).split(",") if x.strip()]

        return [parent] + nested

    if "(" in skill and ")" not in skill:
        return [skill.strip()]   

    if "," in skill:
        return [x.strip() for x in skill.split(",") if x.strip()]

    return [skill]

def get_all_role_skills(role_data):
    skills = []

    for category in ["required", "preferred", "tools"]:
        for skill in role_data.get(category, []):
            skills.extend(expand_skill(skill))

    return list(dict.fromkeys(skills))

def normalize_for_match(s):
    if not s:
        return None
    return normalize_skill(s).lower().strip()

def find_missing_required_skills(user_skills, role_data):

    if not role_data:
        return []

    user_skill_set = {
    normalize_for_match(skill)
    for skill in user_skills
    if normalize_for_match(skill)
    }

    role_skills = get_all_role_skills(role_data)

    missing = []

    for skill in role_skills:

        normalized = normalize_skill(skill)

        if not normalized:
            continue

        role_skill_norm = normalize_for_match(skill)

        if not role_skill_norm:
            continue

        matched = any(
            role_skill_norm in user or user in role_skill_norm
            for user in user_skill_set
        )

        if not matched:
            missing.append(skill)
            
    return list(dict.fromkeys(missing))
                
def main():
    print("\n Adaptive Learning System\n")

    load_dotenv()
    api_key = os.getenv("GEMINI_API_KEY")

    if not api_key:
        print("Missing API key")
        return

    client = genai.Client(api_key=api_key)

    users = load_users()
    roles = load_roles()

    for u in users:
        print(f"[{u['id']}] {u['name']} → {u['target_role']}")

    choice = int(input("\nSelect user ID: "))
    user = users[choice - 1]

    role_name = match_role(user["target_role"], roles)
    role_data = roles.get(role_name)

    user_skills = flatten_user_skills(user["raw_row"])

    aligned = align_skills(user_skills, role_data)

    filtered_skills = filter_relevant_skills(aligned, role_data)
    missing_required_skills = find_missing_required_skills(
    user_skills,
    role_data
    )
    experience = parse_experience(
        user["raw_row"].get("Experience in Target Role")
    )
    difficulty = build_difficulty_map(experience)

    print("\nSelected:", user["name"])
    print("Role:", role_name)
    print("Skills:", filtered_skills)

    result = generate_mcqs(
        client,
        {"skills": filtered_skills, "difficulty": difficulty},
        role_name
    )

    if not result:
        print("API failed")
        return

    for q in result.get("questions", []):
        print(f"\n{q['skill']} ({q['difficulty']})")
        print(q["question"])
        for k, v in q["options"].items():
            print(f"{k}. {v}")
        q['user answer']= input(
            "Your answer: "
        )
    #####################################################
    # TO DO: Store assessment questions and answers in DB 
    #####################################################
    
    with open("results.json", "w") as f:
        json.dump(result, f, indent=2)
    print("\nDone → results.json saved")    
    skill_gap_computation = skill_gap(result)  
    
    ############################################
    # TO DO: Store skill gap, weak skills in DB 
    ############################################
    
    with open("weak_skills.json", "w") as f:
        json.dump(skill_gap_computation, f, indent=2)
    #pathway_prompt = content_generation(user)
    #pathway=call_gemini(client, pathway_prompt)
    #print("*************************** PATHWAY *************************************")
    #print("timeline:", pathway['timeline_weeks'], 'weeks')
    #for content in pathway['learning_path']:
    #    print("SKILL :-  \n",content['skill'])
    #    print("REASON : - \n",content['reason'])
    #    print("RESOURCE SUMMARY - \n", content['resource_summary'])
    #    print("WEEKLY PLAN \n")
    #    for week in content['weekly_plan']:
    #        print("WEEK -", week['week'])
    #        print("FOCUS - ", week['focus'])
    #        print("TASKS - \n", week['TASK'])



if __name__ == "__main__":
    main()
